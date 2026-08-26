"""변호사님의 범죄구성요건 엑셀 → ``knowledge/criminal/*.jsonl``.

엑셀이 원본이고 JSONL은 봇이 읽는 사본입니다. 엑셀을 고치신 뒤 이 명령을
다시 돌리면 됩니다.

    python -m kakao_legal_bot.app.criminal_import 형사데이터.xlsx

변환 규칙 — 모르는 것을 아는 척하지 않는 것이 전부입니다.

* ``미수·예비`` 에 조문이 적혀 있으면(``제342조 미수``) 그 조문을 근거로 처벌,
  ``미수 없음`` 이면 불처벌, ``일반칙``·``검토``·``확인`` 이면 **모름(null)** —
  봇은 "확인 필요"라고 답합니다.
* ``소추조건·특례`` 가 ``친고죄``/``반의사불벌`` 이면 그 값으로, ``해당 없음``
  이면 둘 다 아님으로, 그 밖(친족상도례 검토 등)이면 **원문 그대로** 특례
  필드에 남깁니다.
* ``법정형/확인사항`` 의 "…원문 확인" 은 법정형이 아니라 지시문이므로
  법정형을 비워 두고 출처 URL만 남깁니다.
* 기존 JSONL(씨앗)에 있던 **별칭·질문항목** 은 죄명이 같으면 이어받습니다.
  일상어 매칭("훔쳤어요"→절도)은 손으로 만든 별칭이 필요하니까요.
* ``검증등급`` A → 검증완료, 그 외(B·C) → 미검증. 미검증이면 봇이 답변에
  "변호사 확인 필요"를 붙입니다.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from .criminal import CRIMINAL_DIR

CRIMES_SHEET = "범죄구성요건"
SCHOOL_SHEET = "학교폭력_제재"

_ARTICLE_IN_TEXT = re.compile(r"제\s*(\d{1,4})\s*조(?:\s*의\s*(\d{1,3}))?")
_UNKNOWN_WORDS = ("검토", "확인", "일반칙", "관계", "구조", "포함")


def _clean(value: Any) -> str:
    text = str(value).strip() if value is not None else ""
    return "" if text in {"-", "None", "nan"} else re.sub(r"\s+", " ", text)


def _split_items(*values: str) -> list[str]:
    items: list[str] = []
    for value in values:
        for part in re.split(r"\s*[|·;]\s*|\s*/\s*(?=[가-힣])", value or ""):
            part = part.strip(" .")
            if part and part not in items and part not in {"해당 없음", "누구든지"}:
                items.append(part)
    return items


def _attempt_field(raw: str, statute: str) -> dict[str, Any] | None:
    """``미수·예비`` 칸 → 미수 필드. 모르면 null."""
    text = _clean(raw)
    if not text:
        return None
    if re.search(r"미수\s*없음|별도\s*미수\s*없음", text):
        return {"처벌": False, "근거": ""}
    match = _ARTICLE_IN_TEXT.search(text)
    if match and "미수" in text:
        article = f"제{match.group(1)}조" + (f"의{match.group(2)}" if match.group(2) else "")
        return {"처벌": True, "근거": f"{statute} {article}"}
    # "일반칙" · "미수 검토" · "미수 처벌 확인" — 단정할 근거가 없습니다.
    return None


def _prosecution_flags(raw: str) -> tuple[bool | None, bool | None, str]:
    """``소추조건·특례`` → (친고죄, 반의사불벌, 남길 특례 원문)."""
    text = _clean(raw)
    if not text or text == "해당 없음":
        return (False, False, "")
    if "친고죄" in text:
        return (True, False, text if text != "친고죄" else "")
    if "반의사불벌" in text:
        return (False, True, text if text != "반의사불벌" else "")
    return (None, None, text)


def _penalty(raw: str, url: str) -> tuple[str, str]:
    """'…원문 확인' 은 법정형이 아니라 지시문입니다."""
    text = _clean(raw)
    if not text or "확인" in text:
        return "", text
    return text, ""


def _match_key(name: str) -> str:
    """별칭을 이어받을 때 쓰는 느슨한 죄명 키.

    '정보통신망 명예훼손' 과 '정보통신망명예훼손(구법)' 은 같은 죄의 표기
    차이입니다. 띄어쓰기와 괄호 꼬리를 지워 잇습니다.
    """
    text = re.sub(r"[(（][^)）]*[)）]", "", name or "")
    return re.sub(r"[\s·]", "", text)


def _load_alias_table() -> dict[str, list[str]]:
    """``별칭.json`` — 상담자의 일상어를 죄명에 잇는, 변호사님이 관리하는 표."""
    path = CRIMINAL_DIR / "별칭.json"
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return {
        _clean(name): [_clean(alias) for alias in aliases if _clean(alias)]
        for name, aliases in (raw.get("별칭") or {}).items()
        if _clean(name)
    }


def _load_previous() -> dict[str, dict[str, Any]]:
    """기존 JSONL 의 별칭·질문항목을 죄명별로 모은다."""
    previous: dict[str, dict[str, Any]] = {}
    if not CRIMINAL_DIR.is_dir():
        return previous
    for path in sorted(CRIMINAL_DIR.glob("*.jsonl")):
        if path.name.startswith("학교폭력"):
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith(("#", "//")):
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            name = _clean(row.get("죄명"))
            if name:
                previous.setdefault(_match_key(name), row)
    return previous


def convert_crimes(rows: list[tuple], previous: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        (crime_id, statute, article, name, subject, interest, act, result, intent,
         completion, attempt_raw, prosecution_raw, penalty_raw, _normalised,
         indictment, grade, as_of, url) = (_clean(cell) for cell in row[:18])
        if not name or not statute:
            continue
        if name in seen:
            print(f"  ! 죄명 중복 건너뜀: {name} ({crime_id})", file=sys.stderr)
            continue
        seen.add(name)

        old = previous.get(_match_key(name), {})
        alias_table = _load_alias_table()
        aliases: list[str] = list(old.get("별칭") or [])
        for alias in alias_table.get(name, []):
            if alias not in aliases:
                aliases.append(alias)
        complaint, veto, special = _prosecution_flags(prosecution_raw)
        penalty, penalty_note = _penalty(penalty_raw, url)
        objective = _split_items(subject, act, result)

        record: dict[str, Any] = {
            "죄명": name,
            "별칭": aliases,
            "법률": statute,
            "조문": article,
            "법정형": penalty or old.get("법정형") or "",
            "보호법익": interest,
            "객관적_구성요건": objective,
            "주관적_구성요건": _split_items(intent),
            "기수시기": completion,
            "미수": _attempt_field(attempt_raw, statute)
            if _attempt_field(attempt_raw, statute) is not None
            else old.get("미수"),
            "예비음모": old.get("예비음모"),
            "상습범": old.get("상습범"),
            "과실범": old.get("과실범"),
            "친고죄": complaint if complaint is not None else old.get("친고죄"),
            "반의사불벌죄": veto if veto is not None else old.get("반의사불벌죄"),
            "소추조건_특례": special,
            "공소시효": old.get("공소시효") or "",
            "질문항목": old.get("질문항목") or [],
            "공소사실_기재례": indictment,
            "법정형_확인": penalty_note,
            "출처": url,
            "기준일": as_of,
            "검증": "검증완료" if grade.upper() == "A" else "미검증",
            "검증등급": grade.upper() or "B",
            "id": crime_id,
        }
        out.append({key: value for key, value in record.items() if value not in (None, "", [])})
    return out


def convert_school(rows: list[tuple]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        (category, number, name, content, nature, record_effect, remedy, as_of, url) = (
            _clean(cell) for cell in row[:9]
        )
        if not name:
            continue
        out.append(
            {
                key: value
                for key, value in {
                    "구분": category,
                    "조치번호": number if number != "-" else "",
                    "조치명": name,
                    "내용": content,
                    "법적성격": nature,
                    "학생부_장기효과": record_effect,
                    "불복_병행_확인": remedy,
                    "기준일": as_of,
                    "출처": url,
                }.items()
                if value
            }
        )
    return out


def main(argv: list[str] | None = None) -> int:
    import argparse

    try:
        import openpyxl
    except ImportError:
        print("openpyxl 이 필요합니다: pip install openpyxl", file=sys.stderr)
        return 1

    parser = argparse.ArgumentParser(
        prog="python -m kakao_legal_bot.app.criminal_import",
        description="범죄구성요건 엑셀을 knowledge/criminal/*.jsonl 로 변환",
    )
    parser.add_argument("xlsx", help="변호사님이 관리하는 엑셀 파일")
    parser.add_argument("--out", default=str(CRIMINAL_DIR))
    args = parser.parse_args(argv)

    workbook = openpyxl.load_workbook(args.xlsx, data_only=True)
    if CRIMES_SHEET not in workbook.sheetnames:
        print(f"'{CRIMES_SHEET}' 시트가 없습니다.", file=sys.stderr)
        return 1

    previous = _load_previous()
    crimes = convert_crimes(
        list(workbook[CRIMES_SHEET].iter_rows(min_row=2, values_only=True)), previous
    )
    school = (
        convert_school(list(workbook[SCHOOL_SHEET].iter_rows(min_row=2, values_only=True)))
        if SCHOOL_SHEET in workbook.sheetnames
        else []
    )

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    def dump(path: Path, header: list[str], rows: list[dict[str, Any]]) -> None:
        lines = [f"# {line}" for line in header]
        lines.extend(json.dumps(row, ensure_ascii=False) for row in rows)
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    criminal_code = [row for row in crimes if row["법률"] == "형법"]
    special = [row for row in crimes if row["법률"] != "형법"]
    aliased = sum(1 for row in crimes if row.get("별칭"))

    dump(
        out_dir / "형법각론.jsonl",
        [
            "형법각론 구성요건 — 변호사님 엑셀에서 변환한 파일입니다. 직접 고치지 말고",
            "엑셀을 고친 뒤 python -m kakao_legal_bot.app.criminal_import 로 다시 만드세요.",
            "미수/예비음모/상습범/과실범: 값이 없으면 봇은 '확인 필요'라고 답합니다.",
        ],
        criminal_code,
    )
    dump(
        out_dir / "특별형법.jsonl",
        [
            "특별형법 구성요건 — 형법각론.jsonl 과 같은 방식으로 엑셀에서 변환됩니다.",
        ],
        special,
    )
    if school:
        dump(
            out_dir / "학교폭력.jsonl",
            [
                "학교폭력예방법 제재·조치 — 범죄가 아니라 행정조치입니다.",
                "가해학생 조치(제1~9호) · 피해학생 보호조치 · 절차·불복 항목이 섞여 있습니다.",
            ],
            school,
        )

    print(f"죄명 {len(crimes)}개 (형법 {len(criminal_code)} · 특별형법 {len(special)}) → {out_dir}")
    print(f"별칭 이어받음 {aliased}개 · 학교폭력 항목 {len(school)}개")
    print("점검: python -m kakao_legal_bot.app.criminal --check")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
