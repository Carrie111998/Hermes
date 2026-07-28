#!/usr/bin/env python3
"""Build medic dashboard: JSON data + HTML embedding"""
import openpyxl, json, os

SRC = '/home/angkolj/Desktop/MEDIC FOR HIRE/malaysia-offshore-medic-jobs.xlsx'
OUTPUT_DIR = '/home/angkolj/.hermes/hermes-agent/apps/desktop/medic-job-board'
OUT_JSON = os.path.join(OUTPUT_DIR, 'medic-jobs-data.js')

wb = openpyxl.load_workbook(SRC)
data = {}
for name in wb.sheetnames:
    if 'HOSPITAL' in name.upper():
        display = 'HOSPITAL & CLINIC'
    else:
        display = name
    ws = wb[name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        job = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                h = headers[i].strip()
                if val is not None:
                    if h in ('Phone', 'Email') and not str(val).strip():
                        continue
                    val_str = str(val).strip()
                    if h == 'Remarks' and len(val_str) > 600:
                        val_str = val_str[:600] + '...'
                    if val_str:
                        job[h] = val_str
        if job and 'Job Position' in job:
            jobs.append(job)

    data.setdefault(display, []).extend(jobs)

# Write as JS global variable assignment (loadable via script tag)
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    f.write('// Medic jobs data - auto-generated\n')
    f.write('window.MEDIC_JOBS_DATA = ')
    json.dump(data, f, ensure_ascii=False, indent=1)
    f.write(';\n')

total = sum(len(v) for v in data.values())
f_size = os.path.getsize(OUT_JSON)
print(f'{OUT_JSON}: {total} jobs, {f_size/1024:.0f} KB')

# Verify we can read it back
with open(OUT_JSON, 'r') as f:
    raw = f.read()
    assert 'window.MEDIC_JOBS_DATA' in raw
    print('  - JS variable declaration verified')
